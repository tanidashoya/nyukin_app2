
from flask import Flask, render_template, request
import openpyxl
from datetime import datetime
import math

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # 入力の取得
        tedori_sho = int(request.form["tedori_sho"])
        tedori_ai = int(request.form["tedori_ai"])
        nisa_sho = int(request.form["nisa_sho"])
        nisa_ai = int(request.form["nisa_ai"])
        ideco_sho = int(request.form["ideco_sho"])
        ideco_ai = int(request.form["ideco_ai"])
        denki_keitai_ai = int(request.form["denki_keitai_ai"])
        momiji_nyukin = int(request.form["momiji"])

        # 計算
        tousihiki_sho = tedori_sho - nisa_sho - ideco_sho
        tousihiki_ai = tedori_ai - nisa_ai - ideco_ai
        tedori_result = tousihiki_sho + tousihiki_ai
        kozukai_money = int((tedori_result - momiji_nyukin) / 2)
        nyukin_sho = math.ceil((tousihiki_sho - kozukai_money) / 1000) * 1000
        nyukin_ai = math.ceil((tousihiki_ai - kozukai_money - denki_keitai_ai) / 1000) * 1000
        kozukai_sho = tousihiki_sho - nyukin_sho
        kozukai_ai = tousihiki_ai - nyukin_ai - denki_keitai_ai

        # Excelに保存
        #wb = openpyxl.load_workbook("D:/手取り・入金金額表.xlsx")
        #ws = wb["入力シート"]
        #row = ws.max_row + 1
        #now = datetime.now().strftime("%Y/%m/%d")
        #ws.append([now, tedori_sho, nyukin_sho, kozukai_sho,
        #           tedori_ai, nyukin_ai, kozukai_ai, denki_keitai_ai])
        #wb.save("D:/手取り・入金金額表.xlsx")

        return render_template("index.html", result={
            "nyukin1": nyukin_sho,
            "nyukin2": nyukin_ai,
            "kozukai1": kozukai_sho,
            "kozukai2": kozukai_ai
        })

    return render_template("index.html", result=None)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

from flask import Flask, render_template, request
import openpyxl
from datetime import datetime
import math

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # 入力の取得
        tedori_sho = int(request.form["tedori_sho"])
        tedori_ai = int(request.form["tedori_ai"])
        nisa_sho = int(request.form["nisa_sho"])
        nisa_ai = int(request.form["nisa_ai"])
        ideco_sho = int(request.form["ideco_sho"])
        ideco_ai = int(request.form["ideco_ai"])
        denki_keitai_ai = int(request.form["denki_keitai_ai"])
        momiji_nyukin = int(request.form["momiji"])

        # 計算
        tousihiki_sho = tedori_sho - nisa_sho - ideco_sho
        tousihiki_ai = tedori_ai - nisa_ai - ideco_ai
        tedori_result = tousihiki_sho + tousihiki_ai
        kozukai_money = int((tedori_result - momiji_nyukin) / 2)
        nyukin_sho = math.ceil((tousihiki_sho - kozukai_money) / 1000) * 1000
        nyukin_ai = math.ceil((tousihiki_ai - kozukai_money - denki_keitai_ai) / 1000) * 1000
        kozukai_sho = tousihiki_sho - nyukin_sho
        kozukai_ai = tousihiki_ai - nyukin_ai - denki_keitai_ai

        # Excelに保存
        #wb = openpyxl.load_workbook("D:/手取り・入金金額表.xlsx")
        #ws = wb["入力シート"]
        #row = ws.max_row + 1
        #now = datetime.now().strftime("%Y/%m/%d")
        #ws.append([now, tedori_sho, nyukin_sho, kozukai_sho,
        #           tedori_ai, nyukin_ai, kozukai_ai, denki_keitai_ai])
        #wb.save("D:/手取り・入金金額表.xlsx")

        return render_template("index.html", result={
            "nyukin1": nyukin_sho,
            "nyukin2": nyukin_ai,
            "kozukai1": kozukai_sho,
            "kozukai2": kozukai_ai
        })

    return render_template("index.html", result=None)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
